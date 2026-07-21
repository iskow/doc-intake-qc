#!/usr/bin/env python3
"""Phase 3 — AI classification layer for the Document Intake QC Agent.

Reads manifest.csv, extracts each file's text, and asks a LOCAL language model
what kind of document it is: invoice, contract, correspondence, report, or
other. Writes classifications.csv — one row per file.

Nothing leaves this machine. The model runs under Ollama on 127.0.0.1, so a
client's documents are never uploaded to a third-party API. That is the point
of the design, not a side effect (see DECISIONS 2026-07-21).

Two things this script deliberately does NOT do:

1. It never sends the filename or the folder to the model. `Invoices/INV-2026-
   001.pdf` would hand the model the answer. Keeping the prompt to content only
   means the folder stays INDEPENDENT evidence — Phase 2's rules engine can
   then flag documents whose content disagrees with where they were filed.

2. It never guesses. A file with no readable text (a photo, the archive, an
   empty file) is recorded as `unclassified`, not shoehorned into a label.

Read-only: files under mock-intake/ are opened for reading only. This script
writes exactly one file — classifications.csv at the project root.

Run:  py scripts/classify.py        (Ollama must be running)
Output: classifications.csv at the project root.
"""

from __future__ import annotations

import csv
import json
import logging
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

logging.getLogger("pypdf").setLevel(logging.ERROR)   # same expected noise as scan.py

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "manifest.csv"
CLASSIFICATIONS = ROOT / "classifications.csv"

# --- model configuration ---------------------------------------------------
# Ollama's HTTP API on the loopback address. urllib is used instead of the
# `requests` package: this is one plain JSON POST, and the standard library
# already does it — one less dependency to vet and install.
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
MODEL = "gemma4:12b"

# The five labels. `unclassified` is NOT one of them — the model is never
# offered it. It is what WE record when there is nothing to classify, so a
# missing label always traces to our decision, never to the model's guess.
LABELS = ["invoice", "contract", "correspondence", "report", "other"]
UNCLASSIFIED = "unclassified"

# How much text the model sees. Enough to judge a document, short enough to
# keep each call fast — the label lives in the first page, not the last.
TEXT_LIMIT = 3000

# Below this many characters there is nothing meaningful to classify.
MIN_CHARS = 20

# Folder name -> the label that folder implies. Recorded alongside the model's
# answer so Phase 2 can compare the two. NOT sent to the model. Folders absent
# from this table ("Site Photos", "Old Files", the root) give no hint at all,
# which is honest — they carry no class claim.
FOLDER_HINT = {
    "Invoices": "invoice",
    "Contracts": "contract",
    "Correspondence": "correspondence",
    "Reports": "report",
}

FIELDS = ["path", "name", "chars", "label", "confidence", "folder_hint", "status", "reason"]

PROMPT = """You are classifying a business document for a legal intake review.

Read the document text below and choose exactly one label:
- invoice: a bill or request for payment (invoice number, amount due, terms)
- contract: an agreement between parties (agreement, NDA, MSA, amendment)
- correspondence: a message between people (email, letter, memo)
- report: an analysis or summary of activity over a period
- other: none of the above

Respond with JSON only: {{"label": "<one label>", "confidence": <0.0-1.0>, "reason": "<under 15 words>"}}

Document text:
---
{text}
---"""


def extract_text(path: Path, true_type: str) -> str:
    """Pull readable text out of a file, dispatching on its REAL type.

    Why true_type (the magic-byte class from scan.py) and not the extension:
    the fixture contains `report_final.pdf`, which is plain text wearing a .pdf
    name. Dispatching on the extension would hand it to the PDF reader, which
    fails; dispatching on content reads it correctly.

    Why there is no catch-all "just read it as text" fallback: an earlier draft
    had one, and it silently scraped 541 characters of binary garbage out of a
    PNG and fed it to the model, which then confidently labeled it. A fallback
    that always returns something turns a loud failure into a quiet wrong
    answer. Formats with no text layer return "" here on purpose — images,
    archives, legacy OLE files, and empty files have nothing to read, and
    saying so is the correct result.
    """
    ext = path.suffix.lower()
    try:
        if true_type == "pdf":
            return "\n".join(p.extract_text() or "" for p in PdfReader(str(path)).pages)[:TEXT_LIMIT]
        if true_type == "text":
            # Covers .txt and .csv, plus the mislabeled .pdf and .doc files.
            return path.read_text(encoding="utf-8", errors="strict")[:TEXT_LIMIT]
        if true_type == "zip":
            # OOXML documents are zip containers. Only .docx/.xlsx carry text we
            # can read; a real .zip archive is left alone — the pipeline never
            # silently expands archives (same rule as scan.py).
            if ext == ".docx":
                return "\n".join(p.text for p in Document(str(path)).paragraphs)[:TEXT_LIMIT]
            if ext == ".xlsx":
                ws = load_workbook(str(path), read_only=True).active
                return "\n".join(
                    " | ".join("" if c is None else str(c) for c in row)
                    for row in ws.iter_rows(values_only=True)
                )[:TEXT_LIMIT]
    except Exception:
        # Corrupt or mislabeled file that will not parse as its real type.
        # No text is the honest answer; the caller records it as unclassified.
        return ""
    return ""


def ask_model(text: str) -> dict:
    """Send one document's text to the local model and return its JSON answer.

    temperature=0 makes the answer repeatable — the same document classifies
    the same way on every run, which a QC gate depends on. format="json" makes
    Ollama constrain the output to valid JSON, so we never have to scrape a
    label out of a sentence.
    """
    body = json.dumps({
        "model": MODEL,
        "prompt": PROMPT.format(text=text),
        "stream": False,
        "format": "json",
        "options": {"temperature": 0},
    }).encode()
    req = urllib.request.Request(OLLAMA_URL, data=body,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=180) as response:
        return json.loads(json.loads(response.read())["response"])


def classify_row(rec: dict) -> dict:
    """Build one classifications.csv row for one manifest row.

    status is the honest account of what happened:
      ok           the model returned a valid label
      no_text      nothing readable to classify — never guessed
      model_error  the model failed or answered off-list; surfaced, not hidden
    """
    path = ROOT / rec["path"]
    folder_hint = FOLDER_HINT.get(Path(rec["path"]).parts[1], "") \
        if len(Path(rec["path"]).parts) > 2 else ""

    text = extract_text(path, rec["true_type"])
    row = {
        "path": rec["path"], "name": rec["name"], "chars": len(text),
        "label": UNCLASSIFIED, "confidence": 0.0, "folder_hint": folder_hint,
        "status": "no_text", "reason": "",
    }

    if len(text.strip()) < MIN_CHARS:
        row["reason"] = f"no extractable text ({len(text.strip())} chars)"
        return row

    try:
        answer = ask_model(text)
    except urllib.error.URLError as e:
        # Ollama not running is an operator problem, not a data problem — stop
        # rather than write a file full of errors that looks like a real result.
        raise SystemExit(
            f"Cannot reach Ollama at {OLLAMA_URL} ({e}).\n"
            "Start it with `ollama serve`, then rerun this script."
        ) from e
    except Exception as e:
        row["status"] = "model_error"
        row["reason"] = f"model call failed: {type(e).__name__}"
        return row

    label = str(answer.get("label", "")).lower().strip()
    if label not in LABELS:
        # An off-list answer is a model failure. Recording it as `other` would
        # bury the failure inside a real label.
        row["status"] = "model_error"
        row["reason"] = f"model returned invalid label {label!r}"
        return row

    try:
        confidence = round(float(answer.get("confidence", 0)), 2)
    except (TypeError, ValueError):
        confidence = 0.0

    row.update({
        "label": label, "confidence": confidence, "status": "ok",
        "reason": str(answer.get("reason", "")).strip()[:120],
    })
    return row


def main() -> None:
    with MANIFEST.open(encoding="utf-8") as f:
        records = list(csv.DictReader(f))

    rows = []
    t0 = time.time()
    for i, rec in enumerate(records, 1):
        row = classify_row(rec)
        rows.append(row)
        print(f"  [{i:>2}/{len(records)}] {row['label']:<15} {rec['name'][:52]}",
              flush=True)

    with CLASSIFICATIONS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    elapsed = time.time() - t0
    tally: dict[str, int] = {}
    for row in rows:
        tally[row["label"]] = tally.get(row["label"], 0) + 1

    print(f"\nClassified {len(rows)} files in {elapsed:.1f}s "
          f"({elapsed / max(1, len(rows)):.1f}s each) "
          f"-> {CLASSIFICATIONS.relative_to(ROOT).as_posix()}")
    for label, count in sorted(tally.items()):
        print(f"  {label:<16} {count}")
    errors = sum(1 for r in rows if r["status"] == "model_error")
    if errors:
        print(f"  model errors: {errors}", file=sys.stderr)


if __name__ == "__main__":
    main()
